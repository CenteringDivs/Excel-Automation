
import openpyxl
from openpyxl import Workbook, load_workbook
from bs4 import BeautifulSoup

my_file = load_workbook('excel_report.xlsx')
sheet = my_file.active


# List containing relevant data
product_groups = []

# Loop through the sheet
for i in range(1, sheet.max_row+1):
    if i == 1: # ignore first row
            continue
    if i > 4: # only focus on top 3 product groups (rows 2, 3, and 4)
        break

    current_row = []

    for j in range(1, sheet.max_column+1):
        
        if j == 2:
            continue # ignore second column

        cell_obj = sheet.cell(row=i, column=j).value

        if j == 4 or j == 5:
             # strip whitespace, convert to number
             cell_obj_num = float(str(cell_obj).strip())
             current_row.append(cell_obj_num * 100) # multiply by 100 to format the percentage

        else:
            current_row.append(cell_obj)

    if current_row:
        product_groups.append(current_row)
        

print(product_groups)


with open('index.html', 'r', encoding='utf-8') as file:
    content = file.read()

soup = BeautifulSoup(content, 'html.parser')

# create new html with data from product_groups
row_ids = ['first_row', 'second_row', 'third_row']

for i in range(0, len(product_groups)):
        
        tr_tag = soup.find('tr', id=row_ids[i])
        tr_tag.contents = []

        for j in range(0,4): 
             new_td = soup.new_tag('td')
             new_td.string = str(product_groups[i][j])
             tr_tag.append(new_td)


with open('index.html', 'w', encoding='utf-8') as file:
    file.write(str(soup))
